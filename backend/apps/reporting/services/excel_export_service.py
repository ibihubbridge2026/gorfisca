"""
Service d'export Excel pour les rapports financiers
Génération de fichiers Excel conformes OHADA
"""
import io
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Optional

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter


class ExcelExportService:
    """
    Service d'export des données financières en Excel
    """
    
    def __init__(self):
        # Styles communs
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.cell_font = Font(size=10)
        self.cell_alignment = Alignment(horizontal="left", vertical="center")
        self.amount_alignment = Alignment(horizontal="right", vertical="center")
        
        self.title_font = Font(bold=True, size=14, color="1a1a2e")
        
        # Bordures
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_balance_sheet(self, organization, fiscal_year: int, data: Dict) -> HttpResponse:
        """
        Exporte le bilan comptable en Excel
        
        Args:
            organization: Organisation
            fiscal_year: Année fiscale
            data: Données du bilan structurées
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Bilan"
        
        # En-tête
        ws.merge_cells('A1:E1')
        ws['A1'] = "BILAN COMPTABLE"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Exercice {fiscal_year}"
        ws['A2'].font = Font(size=11, italic=True)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:E3')
        ws['A3'] = f"{organization.name}"
        ws['A3'].font = Font(size=11)
        ws['A3'].alignment = Alignment(horizontal="center")
        
        if hasattr(organization, 'legal_identifier') and organization.legal_identifier:
            ws.merge_cells('A4:E4')
            ws['A4'] = f"NIF: {organization.legal_identifier}"
            ws['A4'].font = Font(size=10)
            ws['A4'].alignment = Alignment(horizontal="center")
        
        # ACTIF
        ws.append([])  # Ligne vide
        ws.append(["ACTIF"])
        ws[f'A{ws.max_row}'].font = self.title_font
        
        # En-têtes de tableau Actif
        headers_actif = ['Code', 'Libellé', 'Brut', 'Amort./Prov.', 'Net']
        ws.append(headers_actif)
        row_num = ws.max_row
        
        # Style des en-têtes
        for col in range(1, len(headers_actif) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données Actif
        for item in data.get('actif', []):
            ws.append([
                item.get('code', ''),
                item.get('label', ''),
                self._format_amount(item.get('brut', 0)),
                self._format_amount(item.get('amortization', 0)),
                self._format_amount(item.get('net', 0))
            ])
            
            # Appliquer bordures et alignement
            for col in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = self.thin_border
                if col >= 3:
                    cell.alignment = self.amount_alignment
                else:
                    cell.alignment = self.cell_alignment
        
        # Ligne vide
        ws.append([])
        
        # PASSIF
        ws.append(["PASSIF"])
        ws[f'A{ws.max_row}'].font = self.title_font
        
        # En-têtes de tableau Passif
        headers_passif = ['Code', 'Libellé', 'Montant']
        ws.append(headers_passif)
        row_num = ws.max_row
        
        for col in range(1, len(headers_passif) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données Passif
        for item in data.get('passif', []):
            ws.append([
                item.get('code', ''),
                item.get('label', ''),
                self._format_amount(item.get('amount', 0))
            ])
            
            for col in range(1, 4):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = self.thin_border
                if col >= 3:
                    cell.alignment = self.amount_alignment
                else:
                    cell.alignment = self.cell_alignment
        
        # Pied de page
        ws.append([])
        ws.append([])
        footer_row = ws.max_row
        ws[f'A{footer_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - GORFISCA v2.0"
        ws[f'A{footer_row}'].font = Font(size=9, italic=True)
        
        # Ajuster largeurs colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # Export
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="bilan_{organization.legal_identifier}_{fiscal_year}.xlsx"'
        )
        response.write(buffer.getvalue())
        
        return response
    
    def export_income_statement(self, organization, fiscal_year: int, data: Dict) -> HttpResponse:
        """
        Exporte le compte de résultat en Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Compte de Résultat"
        
        # En-tête
        ws.merge_cells('A1:C1')
        ws['A1'] = "COMPTE DE RÉSULTAT"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:C2')
        ws['A2'] = f"Exercice {fiscal_year}"
        ws['A2'].font = Font(size=11, italic=True)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:C3')
        ws['A3'] = f"{organization.name}"
        ws['A3'].font = Font(size=11)
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # En-têtes
        headers = ['Rubrique', 'Montant N', 'Montant N-1']
        ws.append([])
        ws.append(headers)
        row_num = ws.max_row
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        for item in data.get('lines', []):
            ws.append([
                item.get('label', ''),
                self._format_amount(item.get('amount_n', 0)),
                self._format_amount(item.get('amount_n_1', 0))
            ])
            
            for col in range(1, 4):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = self.thin_border
                if col >= 2:
                    cell.alignment = self.amount_alignment
                else:
                    cell.alignment = self.cell_alignment
        
        # Résultat net
        net_result = data.get('net_result', 0)
        ws.append([])
        ws.append(['RÉSULTAT NET', self._format_amount(abs(net_result)), ''])
        result_row = ws.max_row
        ws[f'A{result_row}'].font = Font(bold=True, size=12)
        ws[f'B{result_row}'].font = Font(bold=True, size=12)
        if net_result >= 0:
            ws[f'B{result_row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws[f'B{result_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Pied de page
        ws.append([])
        footer_row = ws.max_row
        ws[f'A{footer_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - GORFISCA v2.0"
        ws[f'A{footer_row}'].font = Font(size=9, italic=True)
        
        # Largeurs
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        # Export
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="cr_{organization.legal_identifier}_{fiscal_year}.xlsx"'
        )
        response.write(buffer.getvalue())
        
        return response
    
    def export_general_ledger(self, organization, account_code: str, data: List[Dict]) -> HttpResponse:
        """
        Exporte le grand livre en Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"Grand Livre {account_code}"
        
        # En-tête
        ws.merge_cells('A1:F1')
        ws['A1'] = "GRAND LIVRE"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Compte: {account_code}"
        ws['A2'].font = Font(size=11, italic=True)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"{organization.name}"
        ws['A3'].font = Font(size=11)
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # En-têtes
        headers = ['Date', 'Référence', 'Libellé', 'Débit', 'Crédit', 'Solde']
        ws.append([])
        ws.append(headers)
        row_num = ws.max_row
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données avec solde cumulé
        running_balance = Decimal('0')
        for entry in data:
            amount = Decimal(str(entry.get('amount', 0)))
            if entry.get('line_type') == 'debit':
                running_balance += amount
                debit = amount
                credit = None
            else:
                running_balance -= amount
                debit = None
                credit = amount
            
            ws.append([
                entry.get('date', ''),
                entry.get('reference', ''),
                entry.get('description', '')[:50],
                self._format_amount(debit) if debit else '',
                self._format_amount(credit) if credit else '',
                self._format_amount(running_balance)
            ])
            
            for col in range(1, 7):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = self.thin_border
                if col >= 4:
                    cell.alignment = self.amount_alignment
                else:
                    cell.alignment = self.cell_alignment
        
        # Solde final
        ws.append([])
        ws.append(['Solde final', '', '', '', '', self._format_amount(running_balance)])
        final_row = ws.max_row
        ws[f'A{final_row}'].font = Font(bold=True)
        ws[f'F{final_row}'].font = Font(bold=True)
        
        # Pied de page
        ws.append([])
        footer_row = ws.max_row
        ws[f'A{footer_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - GORFISCA v2.0"
        ws[f'A{footer_row}'].font = Font(size=9, italic=True)
        
        # Largeurs
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Export
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="grand_livre_{account_code}.xlsx"'
        )
        response.write(buffer.getvalue())
        
        return response
    
    def export_reconciliation(self, organization, transactions: List[Dict]) -> HttpResponse:
        """
        Exporte les transactions rapprochées en Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions Rapprochées"
        
        # En-tête
        ws.merge_cells('A1:H1')
        ws['A1'] = "TRANSACTIONS RAPPROCHÉES"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"{organization.name}"
        ws['A2'].font = Font(size=11)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes
        headers = [
            'Date Transaction', 'Description', 'Montant', 
            'Date Rapprochement', 'Compte', 'Référence', 
            'Statut', 'Commentaire'
        ]
        ws.append([])
        ws.append(headers)
        row_num = ws.max_row
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        for txn in transactions:
            ws.append([
                txn.get('transaction_date', ''),
                txn.get('description', '')[:40],
                self._format_amount(txn.get('amount', 0)),
                txn.get('reconciled_at', ''),
                txn.get('account_code', ''),
                txn.get('reference', ''),
                txn.get('status', ''),
                txn.get('notes', '')[:30]
            ])
            
            for col in range(1, 9):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = self.thin_border
                if col == 3:
                    cell.alignment = self.amount_alignment
                else:
                    cell.alignment = self.cell_alignment
        
        # Pied de page
        ws.append([])
        footer_row = ws.max_row
        ws[f'A{footer_row}'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - GORFISCA v2.0"
        ws[f'A{footer_row}'].font = Font(size=9, italic=True)
        
        # Largeurs
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['B'].width = 40
        
        # Export
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="transactions_rapprochees_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        )
        response.write(buffer.getvalue())
        
        return response
    
    def _format_amount(self, value) -> str:
        """Formate un montant avec séparateurs de milliers"""
        if value is None or value == '':
            return ''
        try:
            return f"{float(value):,.2f}"
        except (ValueError, TypeError):
            return str(value)
